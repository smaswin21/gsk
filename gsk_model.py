from __future__ import annotations

import json
import math
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import joblib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.ensemble import RandomForestRegressor
from sklearn.linear_model import LinearRegression
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import mean_absolute_error, mean_squared_error
from sklearn.model_selection import KFold, train_test_split
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler
from xgboost import XGBRegressor


DATA_PATH = Path("sales.xlsx")
ARTIFACT_DIR = Path("artifacts")
MODEL_PATH = ARTIFACT_DIR / "gsk_multinomial_model.joblib"
CONFIG_PATH = ARTIFACT_DIR / "gsk_model_config.json"

INDICATIONS = ("a", "b", "c")
TARGET_COLUMNS = ["avg_split_a", "avg_split_b", "avg_split_c"]
FEATURE_COLUMNS = [
    "log_total_6m_sales",
    "sales_cv",
    "total_touchpoints_all",
    "touchpoints_share_a",
    "touchpoints_share_b",
    "total_hcps_all",
    "hcp_share_a",
    "hcp_share_b",
    "tp_per_hcp_a",
    "tp_per_hcp_b",
    "tp_per_hcp_c",
]
USER_INPUT_COLUMNS = [
    "total_6m_sales",
    "sales_cv",
    "total_touchpoints_all",
    "touchpoints_share_a",
    "touchpoints_share_b",
    "total_hcps_all",
    "hcp_share_a",
    "hcp_share_b",
]

EPSILON = 1e-6


@dataclass
class TrainingArtifacts:
    model: Pipeline
    config: dict[str, Any]
    modeling_df: pd.DataFrame
    training_df: pd.DataFrame
    test_df: pd.DataFrame
    cv_predictions: pd.DataFrame
    benchmark_cv_predictions: pd.DataFrame
    holdout_predictions: pd.DataFrame
    benchmark_holdout_predictions: pd.DataFrame
    cv_metrics: pd.DataFrame
    holdout_metrics: pd.DataFrame
    benchmark_cv_metrics: pd.DataFrame
    benchmark_holdout_metrics: pd.DataFrame
    tuning_results: pd.DataFrame


def _slugify(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text


def _forward_fill(values: list[Any]) -> list[Any]:
    filled: list[Any] = []
    last = None
    for value in values:
        if value is not None and str(value).strip() != "":
            last = value
            filled.append(value)
        else:
            filled.append(last)
    return filled


def load_sales_workbook(filepath: str | Path = DATA_PATH) -> pd.DataFrame:
    filepath = Path(filepath)
    workbook = load_workbook(filepath, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))

    header_rows = [_forward_fill(list(row)) for row in rows[:3]]
    column_names: list[str] = []
    for col_idx in range(len(header_rows[0])):
        parts = [
            _slugify(header_rows[level][col_idx])
            for level in range(3)
            if _slugify(header_rows[level][col_idx])
        ]
        name = "_".join(parts) if parts else f"column_{col_idx + 1}"
        column_names.append(name)

    data = pd.DataFrame(rows[3:], columns=column_names)
    data = data.rename(columns={"hospital_id": "hospital_id"})
    return data


def build_modeling_frame(filepath: str | Path = DATA_PATH) -> pd.DataFrame:
    df = load_sales_workbook(filepath).copy()

    sales_cols = [f"monthly_sales_sales_in_units_m{month}" for month in range(1, 7)]
    split_cols = {
        ind: [f"split_sales_by_indication_m{month}_indication_{ind}" for month in range(1, 7)]
        for ind in INDICATIONS
    }
    tp_cols = {
        ind: [f"indication_{ind}_touchpoints_m{month}" for month in range(1, 7)]
        for ind in INDICATIONS
    }
    hcp_cols = {
        ind: [f"indication_{ind}_hcps_m{month}" for month in range(1, 7)]
        for ind in INDICATIONS
    }

    for column in sales_cols:
        df[column] = pd.to_numeric(df[column], errors="coerce")

    for ind in INDICATIONS:
        for column in split_cols[ind] + tp_cols[ind] + hcp_cols[ind]:
            df[column] = pd.to_numeric(df[column], errors="coerce")

    df["total_6m_sales"] = df[sales_cols].sum(axis=1)
    df["log_total_6m_sales"] = np.log(df["total_6m_sales"])
    df["sales_cv"] = df[sales_cols].std(axis=1, ddof=0) / df[sales_cols].mean(axis=1)

    for ind in INDICATIONS:
        df[f"avg_split_{ind}"] = df[split_cols[ind]].mean(axis=1)
        df[f"total_touchpoints_{ind}"] = df[tp_cols[ind]].sum(axis=1)
        df[f"total_hcps_{ind}"] = df[hcp_cols[ind]].sum(axis=1)

    df["total_touchpoints_all"] = df[[f"total_touchpoints_{ind}" for ind in INDICATIONS]].sum(axis=1)
    df["total_hcps_all"] = df[[f"total_hcps_{ind}" for ind in INDICATIONS]].sum(axis=1)

    for ind in INDICATIONS:
        df[f"touchpoints_share_{ind}"] = df[f"total_touchpoints_{ind}"] / (
            df["total_touchpoints_all"] + EPSILON
        )
        df[f"hcp_share_{ind}"] = df[f"total_hcps_{ind}"] / (df["total_hcps_all"] + EPSILON)
        df[f"tp_per_hcp_{ind}"] = df[f"total_touchpoints_{ind}"] / (
            df[f"total_hcps_{ind}"] + EPSILON
        )

    sum_check = df[TARGET_COLUMNS].sum(axis=1)
    df["target_sum_error"] = np.abs(sum_check - 1.0)

    return df


def get_feature_frame(df: pd.DataFrame) -> pd.DataFrame:
    return df[FEATURE_COLUMNS].copy()


def get_target_frame(df: pd.DataFrame) -> pd.DataFrame:
    return df[TARGET_COLUMNS].copy()

# ---------------------------------------------------------------------------
# Multinomial Logistic Regression (champion model)
# ---------------------------------------------------------------------------
def build_long_format(
    X_hosp: pd.DataFrame,
    y_hosp: pd.DataFrame,
    total_sales: pd.Series,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for index in X_hosp.index:
        row_features = X_hosp.loc[index].to_dict()
        hospital_sales = float(total_sales.loc[index])
        for class_id, target_col in enumerate(TARGET_COLUMNS):
            share = max(float(y_hosp.loc[index, target_col]), EPSILON)
            rows.append(
                {
                    **row_features,
                    "class_id": class_id,
                    "sample_weight": share * hospital_sales,
                }
            )
    return pd.DataFrame(rows)


def make_multinomial_pipeline(C: float) -> Pipeline:
    return Pipeline(
        [
            ("scaler", StandardScaler()),
            (
                "clf",
                LogisticRegression(
                    solver="lbfgs",
                    C=C,
                    max_iter=5000,
                    random_state=42,
                ),
            ),
        ]
    )


def fit_weighted_multinomial(
    X: pd.DataFrame,
    y: pd.DataFrame,
    total_sales: pd.Series,
    C: float,
) -> Pipeline:
    train_long = build_long_format(X, y, total_sales)
    model = make_multinomial_pipeline(C)
    model.fit(
        train_long[FEATURE_COLUMNS],
        train_long["class_id"],
        clf__sample_weight=train_long["sample_weight"],
    )
    return model


def predict_multinomial(model: Pipeline, X: pd.DataFrame) -> pd.DataFrame:
    predictions = model.predict_proba(X[FEATURE_COLUMNS])
    prediction_df = pd.DataFrame(
        predictions,
        columns=["pred_split_a", "pred_split_b", "pred_split_c"],
        index=X.index,
    )
    prediction_df["pred_sum"] = prediction_df.sum(axis=1)
    return prediction_df

# ---------------------------------------------------------------------------
# ALR OLS Benchmark
# ---------------------------------------------------------------------------

def fit_alr_models(X_train: pd.DataFrame, y_train: pd.DataFrame) -> tuple[Any, Any, StandardScaler]:
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_train[FEATURE_COLUMNS])

    y_clipped = y_train.clip(EPSILON, 1 - EPSILON)
    y_alr_a = np.log(y_clipped["avg_split_a"] / y_clipped["avg_split_c"])
    y_alr_b = np.log(y_clipped["avg_split_b"] / y_clipped["avg_split_c"])

    model_a = LinearRegression().fit(X_scaled, y_alr_a)
    model_b = LinearRegression().fit(X_scaled, y_alr_b)
    return model_a, model_b, scaler


def predict_alr(
    models: tuple[Any, Any, StandardScaler],
    X_test: pd.DataFrame,
) -> pd.DataFrame:
    model_a, model_b, scaler = models
    X_scaled = scaler.transform(X_test[FEATURE_COLUMNS])

    log_ratio_a = model_a.predict(X_scaled)
    log_ratio_b = model_b.predict(X_scaled)
    exp_a = np.exp(log_ratio_a)
    exp_b = np.exp(log_ratio_b)
    denominator = 1 + exp_a + exp_b

    prediction_df = pd.DataFrame(
        {
            "pred_split_a": exp_a / denominator,
            "pred_split_b": exp_b / denominator,
            "pred_split_c": 1 / denominator,
        },
        index=X_test.index,
    )
    prediction_df["pred_sum"] = prediction_df.sum(axis=1)
    return prediction_df

# ---------------------------------------------------------------------------
# Dirichlet Regression
# ---------------------------------------------------------------------------

def _dirichlet_log_likelihood(
    params: np.ndarray,
    X: np.ndarray,
    Y: np.ndarray,
    n_targets: int,
    n_features: int,
) -> float:
    """Negative Dirichlet log-likelihood (minimised by scipy)."""
    betas = params.reshape(n_targets, n_features + 1)
    X_aug = np.column_stack([np.ones(len(X)), X])
    log_alphas = X_aug @ betas.T
    alphas = np.exp(log_alphas)
    alpha0 = alphas.sum(axis=1)

    from scipy.special import gammaln
    ll = (
        gammaln(alpha0)
        - gammaln(alphas).sum(axis=1)
        + ((alphas - 1) * np.log(np.clip(Y, EPSILON, 1 - EPSILON))).sum(axis=1)
    ).sum()
    return -ll


def _dirichlet_gradient(
    params: np.ndarray,
    X: np.ndarray,
    Y: np.ndarray,
    n_targets: int,
    n_features: int,
) -> np.ndarray:
    """Gradient of negative Dirichlet log-likelihood."""
    from scipy.special import digamma

    betas = params.reshape(n_targets, n_features + 1)
    X_aug = np.column_stack([np.ones(len(X)), X])
    log_alphas = X_aug @ betas.T
    alphas = np.exp(log_alphas)
    alpha0 = alphas.sum(axis=1)

    psi_alpha0 = digamma(alpha0)
    psi_alphas = digamma(alphas)
    log_y = np.log(np.clip(Y, EPSILON, 1 - EPSILON))

    d_log_alpha = alphas * (psi_alpha0[:, None] - psi_alphas + log_y)
    grad_betas = -(X_aug.T @ d_log_alpha).T
    return grad_betas.ravel()


def fit_dirichlet_model(
    X_train: pd.DataFrame,
    y_train: pd.DataFrame,
) -> tuple[np.ndarray, StandardScaler, int, int]:
    from scipy.optimize import minimize

    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_train[FEATURE_COLUMNS])
    Y = y_train[TARGET_COLUMNS].values.astype(float)
    Y = Y / Y.sum(axis=1, keepdims=True)

    n_features = X_scaled.shape[1]
    n_targets = Y.shape[1]
    params0 = np.zeros(n_targets * (n_features + 1))

    result = minimize(
        _dirichlet_log_likelihood,
        params0,
        jac=_dirichlet_gradient,
        args=(X_scaled, Y, n_targets, n_features),
        method="L-BFGS-B",
        options={"maxiter": 2000, "ftol": 1e-12},
    )

    return result.x, scaler, n_targets, n_features


def predict_dirichlet(
    model_tuple: tuple[np.ndarray, StandardScaler, int, int],
    X_test: pd.DataFrame,
) -> pd.DataFrame:
    betas, scaler, n_targets, n_features = model_tuple
    X_scaled = scaler.transform(X_test[FEATURE_COLUMNS])
    X_aug = np.column_stack([np.ones(len(X_scaled)), X_scaled])

    betas_mat = betas.reshape(n_targets, n_features + 1)
    alphas = np.exp(X_aug @ betas_mat.T)
    shares = alphas / alphas.sum(axis=1, keepdims=True)

    prediction_df = pd.DataFrame(
        shares,
        columns=["pred_split_a", "pred_split_b", "pred_split_c"],
        index=X_test.index,
    )
    prediction_df["pred_sum"] = prediction_df.sum(axis=1)
    return prediction_df


# ---------------------------------------------------------------------------
# Random Forest (3 regressors + normalisation)
# ---------------------------------------------------------------------------

def fit_random_forest_models(
    X_train: pd.DataFrame,
    y_train: pd.DataFrame,
    n_estimators: int = 300,
    random_state: int = 42,
) -> tuple[RandomForestRegressor, RandomForestRegressor, RandomForestRegressor, StandardScaler]:
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_train[FEATURE_COLUMNS])

    rf_a = RandomForestRegressor(n_estimators=n_estimators, random_state=random_state, n_jobs=-1)
    rf_b = RandomForestRegressor(n_estimators=n_estimators, random_state=random_state, n_jobs=-1)
    rf_c = RandomForestRegressor(n_estimators=n_estimators, random_state=random_state, n_jobs=-1)

    rf_a.fit(X_scaled, y_train["avg_split_a"])
    rf_b.fit(X_scaled, y_train["avg_split_b"])
    rf_c.fit(X_scaled, y_train["avg_split_c"])

    return rf_a, rf_b, rf_c, scaler


def predict_random_forest(
    models: tuple[RandomForestRegressor, RandomForestRegressor, RandomForestRegressor, StandardScaler],
    X_test: pd.DataFrame,
) -> pd.DataFrame:
    rf_a, rf_b, rf_c, scaler = models
    X_scaled = scaler.transform(X_test[FEATURE_COLUMNS])

    raw = np.column_stack([
        np.clip(rf_a.predict(X_scaled), EPSILON, None),
        np.clip(rf_b.predict(X_scaled), EPSILON, None),
        np.clip(rf_c.predict(X_scaled), EPSILON, None),
    ])
    shares = raw / raw.sum(axis=1, keepdims=True)

    prediction_df = pd.DataFrame(
        shares,
        columns=["pred_split_a", "pred_split_b", "pred_split_c"],
        index=X_test.index,
    )
    prediction_df["pred_sum"] = prediction_df.sum(axis=1)
    return prediction_df


# ---------------------------------------------------------------------------
# XGBoost (3 regressors + normalisation)
# ---------------------------------------------------------------------------

def fit_xgboost_models(
    X_train: pd.DataFrame,
    y_train: pd.DataFrame,
    n_estimators: int = 200,
    learning_rate: float = 0.05,
    max_depth: int = 3,
    random_state: int = 42,
) -> tuple[XGBRegressor, XGBRegressor, XGBRegressor, StandardScaler]:
    scaler = StandardScaler()
    X_scaled = scaler.fit_transform(X_train[FEATURE_COLUMNS])

    common_kwargs = dict(
        n_estimators=n_estimators,
        learning_rate=learning_rate,
        max_depth=max_depth,
        subsample=0.8,
        colsample_bytree=0.8,
        random_state=random_state,
        verbosity=0,
        n_jobs=-1,
    )

    xgb_a = XGBRegressor(**common_kwargs)
    xgb_b = XGBRegressor(**common_kwargs)
    xgb_c = XGBRegressor(**common_kwargs)

    xgb_a.fit(X_scaled, y_train["avg_split_a"])
    xgb_b.fit(X_scaled, y_train["avg_split_b"])
    xgb_c.fit(X_scaled, y_train["avg_split_c"])

    return xgb_a, xgb_b, xgb_c, scaler


def predict_xgboost(
    models: tuple[XGBRegressor, XGBRegressor, XGBRegressor, StandardScaler],
    X_test: pd.DataFrame,
) -> pd.DataFrame:
    xgb_a, xgb_b, xgb_c, scaler = models
    X_scaled = scaler.transform(X_test[FEATURE_COLUMNS])

    raw = np.column_stack([
        np.clip(xgb_a.predict(X_scaled), EPSILON, None),
        np.clip(xgb_b.predict(X_scaled), EPSILON, None),
        np.clip(xgb_c.predict(X_scaled), EPSILON, None),
    ])
    shares = raw / raw.sum(axis=1, keepdims=True)

    prediction_df = pd.DataFrame(
        shares,
        columns=["pred_split_a", "pred_split_b", "pred_split_c"],
        index=X_test.index,
    )
    prediction_df["pred_sum"] = prediction_df.sum(axis=1)
    return prediction_df

def evaluate_predictions(y_true: pd.DataFrame, y_pred: pd.DataFrame, dataset_label: str, tolerance: float = 0.15) -> pd.DataFrame:
    metrics: list[dict[str, Any]] = []
    for target_col, pred_col, label in zip(
        TARGET_COLUMNS,
        ["pred_split_a", "pred_split_b", "pred_split_c"],
        ["A", "B", "C"],
    ):
        errors = np.abs(y_true[target_col].values - y_pred[pred_col].values)
        accuracy = float((errors <= tolerance).mean() * 100)
        metrics.append(
            {
                "dataset": dataset_label,
                "indication": label,
                "mae": float(mean_absolute_error(y_true[target_col], y_pred[pred_col])),
                "rmse": float(
                    math.sqrt(mean_squared_error(y_true[target_col], y_pred[pred_col]))
                ),
                "accuracy_pct": round(accuracy, 1),
            }
        )
    return pd.DataFrame(metrics)


def select_best_c(
    training_df: pd.DataFrame,
    c_grid: tuple[float, ...] = (0.3, 0.7, 1.5),
    n_splits: int = 5,
    random_state: int = 42,
) -> tuple[float, pd.DataFrame, pd.DataFrame]:
    X_train = get_feature_frame(training_df)
    y_train = get_target_frame(training_df)
    sales_train = training_df["total_6m_sales"]

    folds = list(KFold(n_splits=n_splits, shuffle=True, random_state=random_state).split(X_train))
    candidate_rows: list[dict[str, Any]] = []
    best_predictions = None
    best_score = float("inf")
    best_c = c_grid[0]

    for c_value in c_grid:
        fold_predictions = pd.DataFrame(index=training_df.index)
        fold_predictions[TARGET_COLUMNS] = y_train
        for fold_id, (train_idx, valid_idx) in enumerate(folds, start=1):
            train_ids = X_train.index[train_idx]
            valid_ids = X_train.index[valid_idx]

            model = fit_weighted_multinomial(
                X_train.loc[train_ids],
                y_train.loc[train_ids],
                sales_train.loc[train_ids],
                C=c_value,
            )
            predictions = predict_multinomial(model, X_train.loc[valid_ids])
            fold_predictions.loc[valid_ids, ["pred_split_a", "pred_split_b", "pred_split_c", "pred_sum"]] = (
                predictions
            )

            fold_metrics = evaluate_predictions(
                y_train.loc[valid_ids],
                predictions,
                dataset_label=f"cv_fold_{fold_id}",
            )
            candidate_rows.extend(
                {
                    "C": c_value,
                    "fold": fold_id,
                    **metric_row,
                }
                for metric_row in fold_metrics.to_dict(orient="records")
            )

        mean_mae = float(
            np.mean(
                [
                    row["mae"]
                    for row in candidate_rows
                    if row["C"] == c_value
                ]
            )
        )
        if mean_mae < best_score:
            best_score = mean_mae
            best_c = c_value
            best_predictions = fold_predictions.copy()

    tuning_results = pd.DataFrame(candidate_rows)
    if best_predictions is None:
        raise RuntimeError("Cross-validation did not produce predictions.")
    return best_c, tuning_results, best_predictions


def cross_validate_alr(
    training_df: pd.DataFrame,
    n_splits: int = 5,
    random_state: int = 42,
) -> pd.DataFrame:
    X_train = get_feature_frame(training_df)
    y_train = get_target_frame(training_df)
    predictions = pd.DataFrame(index=training_df.index)
    predictions[TARGET_COLUMNS] = y_train

    kfold = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
    for train_idx, valid_idx in kfold.split(X_train):
        train_ids = X_train.index[train_idx]
        valid_ids = X_train.index[valid_idx]
        models = fit_alr_models(X_train.loc[train_ids], y_train.loc[train_ids])
        fold_predictions = predict_alr(models, X_train.loc[valid_ids])
        predictions.loc[valid_ids, ["pred_split_a", "pred_split_b", "pred_split_c", "pred_sum"]] = (
            fold_predictions
        )

    return predictions


def build_prediction_frame(
    base_df: pd.DataFrame,
    predictions: pd.DataFrame,
    model_label: str,
    dataset_label: str,
) -> pd.DataFrame:
    output = base_df[["hospital_id", *TARGET_COLUMNS, "total_6m_sales"]].copy()
    output = output.join(predictions[["pred_split_a", "pred_split_b", "pred_split_c", "pred_sum"]])
    output["model"] = model_label
    output["dataset"] = dataset_label
    return output


def summarize_tuning_results(tuning_results: pd.DataFrame) -> pd.DataFrame:
    return (
        tuning_results.groupby("C", as_index=False)
        .agg(mean_mae=("mae", "mean"), mean_rmse=("rmse", "mean"))
        .sort_values(["mean_mae", "mean_rmse"])
    )


def build_presets(df: pd.DataFrame) -> dict[str, dict[str, float]]:
    quantiles = df["total_6m_sales"].quantile([0.33, 0.66]).tolist()
    bands = pd.cut(
        df["total_6m_sales"],
        bins=[-np.inf, quantiles[0], quantiles[1], np.inf],
        labels=["Small", "Medium", "Large"],
    )

    preset_columns = USER_INPUT_COLUMNS + ["touchpoints_share_c", "hcp_share_c"]
    preset_df = df.assign(size_band=bands)
    presets: dict[str, dict[str, float]] = {}

    for label in ["Small", "Medium", "Large"]:
        slice_df = preset_df.loc[preset_df["size_band"] == label, preset_columns]
        median_row = slice_df.median(numeric_only=True)
        presets[label.lower()] = {
            key: float(median_row[key])
            for key in preset_columns
        }

    return presets


def build_config(
    full_df: pd.DataFrame,
    best_c: float,
    cv_metrics: pd.DataFrame,
    holdout_metrics: pd.DataFrame,
    benchmark_cv_metrics: pd.DataFrame,
    benchmark_holdout_metrics: pd.DataFrame,
    tuning_summary: pd.DataFrame,
    final_model: Pipeline,
) -> dict[str, Any]:
    feature_ranges: dict[str, dict[str, float]] = {}
    for column in USER_INPUT_COLUMNS:
        quantiles = full_df[column].quantile([0.05, 0.5, 0.95])
        feature_ranges[column] = {
            "min": float(full_df[column].min()),
            "p05": float(quantiles.loc[0.05]),
            "median": float(quantiles.loc[0.5]),
            "p95": float(quantiles.loc[0.95]),
            "max": float(full_df[column].max()),
        }

    coefficient_frame = pd.DataFrame(
        final_model.named_steps["clf"].coef_,
        columns=FEATURE_COLUMNS,
        index=["A", "B", "C"],
    )
    top_drivers = {
        indication: {
            "positive": coefficient_frame.loc[indication].sort_values(ascending=False).head(5).round(4).to_dict(),
            "negative": coefficient_frame.loc[indication].sort_values().head(5).round(4).to_dict(),
        }
        for indication in coefficient_frame.index
    }

    return {
        "data_path": str(DATA_PATH),
        "feature_columns": FEATURE_COLUMNS,
        "target_columns": TARGET_COLUMNS,
        "best_c": best_c,
        "feature_ranges": feature_ranges,
        "presets": build_presets(full_df),
        "cv_metrics": cv_metrics.round(4).to_dict(orient="records"),
        "holdout_metrics": holdout_metrics.round(4).to_dict(orient="records"),
        "benchmark_cv_metrics": benchmark_cv_metrics.round(4).to_dict(orient="records"),
        "benchmark_holdout_metrics": benchmark_holdout_metrics.round(4).to_dict(orient="records"),
        "tuning_summary": tuning_summary.round(4).to_dict(orient="records"),
        "top_drivers": top_drivers,
    }


def prepare_app_features_from_inputs(inputs: dict[str, float]) -> pd.DataFrame:
    touchpoints_share_a = inputs["touchpoints_share_a"]
    touchpoints_share_b = inputs["touchpoints_share_b"]
    touchpoints_share_c = max(0.0, 1.0 - touchpoints_share_a - touchpoints_share_b)

    hcp_share_a = inputs["hcp_share_a"]
    hcp_share_b = inputs["hcp_share_b"]
    hcp_share_c = max(0.0, 1.0 - hcp_share_a - hcp_share_b)

    total_touchpoints_all = inputs["total_touchpoints_all"]
    total_hcps_all = inputs["total_hcps_all"]

    total_touchpoints = {
        "a": total_touchpoints_all * touchpoints_share_a,
        "b": total_touchpoints_all * touchpoints_share_b,
        "c": total_touchpoints_all * touchpoints_share_c,
    }
    total_hcps = {
        "a": total_hcps_all * hcp_share_a,
        "b": total_hcps_all * hcp_share_b,
        "c": total_hcps_all * hcp_share_c,
    }

    feature_row = {
        "log_total_6m_sales": math.log(max(inputs["total_6m_sales"], 1.0)),
        "sales_cv": inputs["sales_cv"],
        "total_touchpoints_all": total_touchpoints_all,
        "touchpoints_share_a": touchpoints_share_a,
        "touchpoints_share_b": touchpoints_share_b,
        "total_hcps_all": total_hcps_all,
        "hcp_share_a": hcp_share_a,
        "hcp_share_b": hcp_share_b,
        "tp_per_hcp_a": total_touchpoints["a"] / (total_hcps["a"] + EPSILON),
        "tp_per_hcp_b": total_touchpoints["b"] / (total_hcps["b"] + EPSILON),
        "tp_per_hcp_c": total_touchpoints["c"] / (total_hcps["c"] + EPSILON),
    }
    return pd.DataFrame([feature_row], columns=FEATURE_COLUMNS)


def train_and_export(
    filepath: str | Path = DATA_PATH,
    artifact_dir: str | Path = ARTIFACT_DIR,
) -> TrainingArtifacts:
    full_df = build_modeling_frame(filepath)

    training_df, test_df = train_test_split(
        full_df,
        test_size=0.2,
        random_state=42,
        shuffle=True,
    )
    training_df = training_df.sort_index().copy()
    test_df = test_df.sort_index().copy()

    best_c, tuning_results, cv_predictions = select_best_c(training_df)
    benchmark_cv_predictions = cross_validate_alr(training_df)

    cv_metrics = evaluate_predictions(
        get_target_frame(training_df),
        cv_predictions,
        dataset_label="cross_validation",
    )
    benchmark_cv_metrics = evaluate_predictions(
        get_target_frame(training_df),
        benchmark_cv_predictions,
        dataset_label="cross_validation",
    )

    champion_model = fit_weighted_multinomial(
        get_feature_frame(training_df),
        get_target_frame(training_df),
        training_df["total_6m_sales"],
        C=best_c,
    )
    holdout_predictions = predict_multinomial(champion_model, get_feature_frame(test_df))
    holdout_metrics = evaluate_predictions(
        get_target_frame(test_df),
        holdout_predictions,
        dataset_label="holdout",
    )

    benchmark_models = fit_alr_models(get_feature_frame(training_df), get_target_frame(training_df))
    benchmark_holdout_predictions = predict_alr(benchmark_models, get_feature_frame(test_df))
    benchmark_holdout_metrics = evaluate_predictions(
        get_target_frame(test_df),
        benchmark_holdout_predictions,
        dataset_label="holdout",
    )

    final_model = fit_weighted_multinomial(
        get_feature_frame(full_df),
        get_target_frame(full_df),
        full_df["total_6m_sales"],
        C=best_c,
    )

    tuning_summary = summarize_tuning_results(tuning_results)
    config = build_config(
        full_df=full_df,
        best_c=best_c,
        cv_metrics=cv_metrics,
        holdout_metrics=holdout_metrics,
        benchmark_cv_metrics=benchmark_cv_metrics,
        benchmark_holdout_metrics=benchmark_holdout_metrics,
        tuning_summary=tuning_summary,
        final_model=final_model,
    )

    artifact_dir = Path(artifact_dir)
    artifact_dir.mkdir(parents=True, exist_ok=True)
    joblib.dump(final_model, artifact_dir / MODEL_PATH.name)
    (artifact_dir / CONFIG_PATH.name).write_text(json.dumps(config, indent=2))

    return TrainingArtifacts(
        model=final_model,
        config=config,
        modeling_df=full_df,
        training_df=training_df,
        test_df=test_df,
        cv_predictions=build_prediction_frame(
            training_df,
            cv_predictions,
            model_label="Weighted multinomial logit",
            dataset_label="Cross-validation",
        ),
        benchmark_cv_predictions=build_prediction_frame(
            training_df,
            benchmark_cv_predictions,
            model_label="ALR OLS benchmark",
            dataset_label="Cross-validation",
        ),
        holdout_predictions=build_prediction_frame(
            test_df,
            holdout_predictions,
            model_label="Weighted multinomial logit",
            dataset_label="Holdout",
        ),
        benchmark_holdout_predictions=build_prediction_frame(
            test_df,
            benchmark_holdout_predictions,
            model_label="ALR OLS benchmark",
            dataset_label="Holdout",
        ),
        cv_metrics=cv_metrics,
        holdout_metrics=holdout_metrics,
        benchmark_cv_metrics=benchmark_cv_metrics,
        benchmark_holdout_metrics=benchmark_holdout_metrics,
        tuning_results=tuning_summary,
    )


def load_exported_artifacts(artifact_dir: str | Path = ARTIFACT_DIR) -> tuple[Pipeline, dict[str, Any]]:
    artifact_dir = Path(artifact_dir)
    model = joblib.load(artifact_dir / MODEL_PATH.name)
    config = json.loads((artifact_dir / CONFIG_PATH.name).read_text())
    return model, config


def pretty_metric_table(metrics: pd.DataFrame) -> pd.DataFrame:
    table = metrics.copy()
    table["mae"] = table["mae"].round(4)
    table["rmse"] = table["rmse"].round(4)
    return table


def main() -> None:
    artifacts = train_and_export()
    print("Saved model artifacts to:")
    print(f"- {MODEL_PATH}")
    print(f"- {CONFIG_PATH}")
    print("\nBest regularization strength (C):", artifacts.config["best_c"])
    print("\nCross-validation metrics:")
    print(pretty_metric_table(artifacts.cv_metrics).to_string(index=False))
    print("\nHoldout metrics:")
    print(pretty_metric_table(artifacts.holdout_metrics).to_string(index=False))


if __name__ == "__main__":
    main()
